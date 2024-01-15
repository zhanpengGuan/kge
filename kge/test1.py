import torch
import pandas as pd
import openpyxl
def calculate_average_size(choice, rank_e):
    # 创建一个字典，将rank_e与相应的choice值关联起来
    rank_size_dict = {}
    for i in range(len(choice)):
        index = int(rank_e[i])
        size = choice[i]
        if index not in rank_size_dict:
            rank_size_dict[index] = [size]
        else:
            rank_size_dict[index].append(size)

    # 计算每个排名的平均大小
    average_sizes = {}
    num_sizes = {}
    for index, sizes in rank_size_dict.items():
        average_sizes[index] = sum(sizes) / len(sizes)
        num_sizes[index] = len(sizes)
    # 计算所有排名的平均大小
    overall_average_size = sum(average_sizes.values()) / len(average_sizes)
    # result_df = pd.DataFrame(list(average_sizes.items()), columns=['排名', '平均大小']).sort_values(by='排名')
    # result_df.to_excel('a.xlsx')
    return average_sizes, num_sizes
# 定义你的模型
checkpoint_path = '/home/guanzp/code/AdaE/kge/local/wnrr/auto/20240107-152247AdaE_auto-auto-cie--0.28--0.5-soft-512-drop-0.36small-gumbel/checkpoint_best.pt'
# data_path = '/home/guanzp/code/AdaE/kge/data/fb15k-237'
data_path = '/home/guanzp/code/AdaE/kge/data/wnrr'
# data_path = '/home/guanzp/code/AdaE/kge/data/yago3-10'
rank_e = torch.load(data_path+'/[-1]rank_e.pt')





# 使用PyTorch加载检查点
checkpoint = torch.load(checkpoint_path)
emb = checkpoint['choice_emb']
Gpro =  torch.zeros(emb.shape,device='cuda:3')
choice = torch.argmax(emb, dim = -1)
# 打印参数
# uni_choice = torch.unique(choice)

# for i in uni_choice:
#   print(i,sum(choice==i))
# print(uni_choice)

# 调用函数并输出结果
result, avg = calculate_average_size(choice, rank_e)
for i in range(0,max(rank_e)):
  print(i,int(result[i]),avg[i])
workbook = openpyxl.Workbook()
sheet = workbook.active

# 在 Excel 表中写入标题
sheet['A1'] = 'rank值'
sheet['B1'] = '平均size'
sheet['C1'] = 'nums'

# 将结果写入 Excel 表
for i in range(max(rank_e)):
    sheet.cell(row=i+2, column=1, value=i)
    sheet.cell(row=i+2, column=2, value=int(result[i]))
    sheet.cell(row=i+2, column=3, value=avg[i])

# 将工作簿保存到文件
workbook.save('/home/guanzp/code/AdaE/kge/结果.xlsx')

# 打印消息，指示数据已写入 Excel 文件
print("结果已写入 '结果.xlsx'")